import sys
sys.path.append("C:\\Users\\it-support\\pyman")  
from fastapi import FastAPI, HTTPException, Depends
import asyncpg
import os
from dotenv import load_dotenv
from sqlalchemy.ext.asyncio import create_async_engine
import logging
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from scraper.parsers.rozetka_parser import RozetkaParser
from fastapi import FastAPI, File, UploadFile
from app.routers.prozorro_router import prozorro_router
from statistics import mean
import logging
from io import BytesIO
from openpyxl import load_workbook
from pydantic import BaseModel
from typing import List, Optional
from .prozorro_functionality.prozorro import get_contract_info
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from urllib.parse import quote
import time
import re
from app.routers.ai_assistant import router as assistant_router




# Local imports
from .routers import crud, xpath
from .routers.authorization import router as auth_router, get_current_user
from .database import Base
from .db_test import verify_database_connection
# ... решта коду без змін

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

load_dotenv()
app = FastAPI(
    title="CheckIT API",
    description="API for tender checking and analysis",
    version="1.0.0"
)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173", "http://localhost:5174", "http://localhost:5175", "http://localhost:5176", "http://localhost:5177",
        "http://localhost:5178", "http://localhost:5179", "http://localhost:5180", "http://localhost:5181", "http://localhost:5182",
        "http://127.0.0.1:5173", "http://127.0.0.1:5174", "http://127.0.0.1:5175", "http://127.0.0.1:5176", "http://127.0.0.1:5177",
        "http://127.0.0.1:5178", "http://127.0.0.1:5179", "http://127.0.0.1:5180", "http://127.0.0.1:5181", "http://127.0.0.1:5182"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Database initialization
async def init_db():
    try:
        if not await verify_database_connection():
            raise Exception("Database connection verification failed")

        engine = create_async_engine(
            os.getenv("DATABASE_URL"),
            echo=True,
            pool_pre_ping=True,
            pool_size=5,
            max_overflow=10
        )
        
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)
            logger.info("Database tables created successfully")
            
        return engine
    except Exception as e:
        logger.error(f"Database initialization failed: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Database initialization failed: {str(e)}"
        )

@app.on_event("startup")
async def startup():
    try:
        await init_db()
        logger.info("Application startup completed successfully")
    except Exception as e:
        logger.error(f"Application startup failed: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Application startup failed: {str(e)}"
        )

app.include_router(assistant_router)
# Include routers with authentication
app.include_router(auth_router, prefix="/auth", tags=["Authentication"])
app.include_router(
    crud.router,
    prefix="/api",
    tags=["CRUD Operations"],
    dependencies=[Depends(get_current_user)]
)
app.include_router(
    xpath.router,
    prefix="/xpath",
    tags=["XPath Operations"],
    dependencies=[Depends(get_current_user)]
)
app.include_router(prozorro_router)


@app.get("/", tags=["Root"])
async def home() -> dict:
    return {
        "message": "Welcome to CheckIT API",
        "documentation": "/docs"
    }

async def test_db_connection():
    try:
        conn = await asyncpg.connect(
            user=os.getenv("POSTGRES_USER"),
            password=os.getenv("POSTGRES_PASSWORD"),
            database=os.getenv("POSTGRES_DB"),
            host=os.getenv("POSTGRES_HOST"),
            port=os.getenv("POSTGRES_PORT")
        )
        await conn.execute('SELECT 1')
        await conn.close()
        logger.info("Test database connection successful")
        return {
            "status": "success",
            "message": "Database connection successful",
            "details": {
                "host": os.getenv("POSTGRES_HOST"),
                "database": os.getenv("POSTGRES_DB"),
                "port": os.getenv("POSTGRES_PORT")
            }
        }
    except Exception as e:
        logger.error(f"Test database connection failed: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Database connection failed: {str(e)}"
        )

def parse_excel_file(file_content: bytes):
    try:
        workbook = load_workbook(filename=BytesIO(file_content))
        sheet = workbook.active
        
        headers = [cell.value for cell in sheet[1]]
        required_columns = ['Назва товару', 'Кількість', 'Ціна', 'Сума']
        
        if not all(col in headers for col in required_columns):
            raise ValueError("Excel file is missing required columns")
        
        col_indices = {header: idx for idx, header in enumerate(headers)}
        
        data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            quantity = row[col_indices['Кількість']]
            price = row[col_indices['Ціна']]
            total = row[col_indices['Сума']]
            
            if quantity is None or price is None or total is None:
                logger.info(f"Skipping row {row_idx} due to empty values")
                continue
                
            try:
                row_data = {
                    'product_name': row[col_indices['Назва товару']],
                    'original_quantity': int(quantity),
                    'original_price_uah': float(price),
                    'original_total_uah': float(total),
                    'rozetka_results': []  # Залишаємо для сумісності, але не використовуємо тут
                }
                data.append(row_data)
            except (ValueError, TypeError) as e:
                logger.warning(f"Skipping row {row_idx} due to invalid data: {str(e)}")
                continue
        
        if not data:
            raise ValueError("No valid data found in the Excel file")
            
        return data
        
    except Exception as e:
        raise ValueError(f"Error parsing Excel file: {str(e)}")

@app.post("/excel-page")
async def upload_excel(file: UploadFile = File(...)):
    rozetka = RozetkaParser()
    try:
        file_content = await file.read()
        excel_data = parse_excel_file(file_content)  # Зберігаємо дані з Excel
        
        d = []  # Результати з Rozetka
        
        for row in excel_data:
            product_name = row['product_name']
            logger.info(f"Searching for: '{product_name}'")
            try:
                rozetka_data = rozetka.find_n_products(
                    product=product_name,
                    n=1,
                    fast_parse=False,
                    ignore_price_format=True,
                    raise_exception=True
                )
                d.extend(rozetka_data)
                logger.info(f"Success: Found {len(rozetka_data)} products for '{product_name}'")
            except Exception as e:
                logger.warning(f"No result for '{product_name}': {str(e)}")
                continue
        
        # Повертаємо обидва набори даних
        return {
            "status": "success",
            "message": "Products found",
            "excel_data": excel_data,  # Дані з Excel
            "rozetka_data": d          # Дані з Rozetka
        }
    except Exception as e:
        logger.error(f"Error in upload_excel: {str(e)}")
        return str(e)

# @app.post("/search-tender")
# async def prozorro_data(contract_id : str):
#     rozetka = RozetkaParser()
#     try:
#         pr_data = get_contract_info(contract_id)
#         d = [] 
        
#         for row in pr_data:
#             product_name = row['name']
#             logger.info(f"Searching for: '{product_name}'")
#             try:
#                 rozetka_data = rozetka.find_n_products(
#                     product=product_name,
#                     n=1,
#                     fast_parse=False,
#                     ignore_price_format=True,
#                     raise_exception=False
#                 )
#                 d.extend(rozetka_data)
#                 logger.info(f"Success: Found {len(rozetka_data)} products for '{product_name}'")
#             except Exception as e:
#                 logger.warning(f"No result for '{product_name}': {str(e)}")
#                 continue
        
#         # Повертаємо обидва набори даних
#         return {
#             "status": "success",
#             "message": "Products found",
#             "excel_data": pr_data,  # Дані з Excel
#             "rozetka_data": d          # Дані з Rozetka
#         }
#     except Exception as e:
#         logger.error(f"Error in upload_excel: {str(e)}")
#         return str(e)

